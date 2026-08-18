#!/usr/bin/env python3
"""
R2 Excel schema monitor for sheeel.com scrapers (all repos, single HyperStack runner).

Validates .xlsx files uploaded to Cloudflare R2 against websites-config.yml excel_schema.
Reports and rolling stats are written back to R2 only — never to the repo.
"""

from __future__ import annotations

import argparse
import fnmatch
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from typing import Any

import boto3
import pandas as pd
import yaml
from botocore.config import Config
from botocore.exceptions import ClientError

from ads_counter import count_scraper_ads
from github_workflows import build_scraper_run_meta, load_site_run_meta
from r2_file_counter import (
    count_daily_r2_inventory_by_type,
    count_scraper_r2_inventory,
    count_scraper_r2_inventory_by_type,
    count_site_r2_inventory,
    count_site_r2_inventory_by_type,
    merge_r2_type_inventories,
)
from request_metrics import (
    aggregate_site_request_metrics,
    build_run_error_summary,
    count_scraper_request_metrics,
)
from openpyxl import load_workbook

MONITOR_SUBPATH = "monitor"
DEFAULT_R2_PREFIX = os.environ.get("R2_PREFIX", "sheeel_data")
ROW_COUNT_TOLERANCE = float(os.environ.get("ROW_COUNT_TOLERANCE", "0.15"))
ROW_COUNT_NO_MAX = 9_999_999
CONFIG_R2_KEY = os.environ.get(
    "WEBSITES_CONFIG_R2_KEY",
    f"{DEFAULT_R2_PREFIX}/{MONITOR_SUBPATH}/websites-config.yml",
)
DEFAULT_CATEGORIES_REGISTRY = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "category_monitor",
    "categories.json",
)
CATEGORIES_REGISTRY_R2_KEY = os.environ.get(
    "CATEGORIES_REGISTRY_R2_KEY",
    f"{DEFAULT_R2_PREFIX}/{MONITOR_SUBPATH}/categories.json",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate sheeel.com Excel files in R2")
    parser.add_argument(
        "--date",
        help="End date YYYY-MM-DD (default: yesterday UTC)",
    )
    parser.add_argument(
        "--days-lookback",
        type=int,
        default=1,
        help="Number of days to inspect ending at --date (default: 1)",
    )
    parser.add_argument(
        "--update-stats",
        action="store_true",
        help="Merge observations into monitor_stats.yml in R2",
    )
    parser.add_argument(
        "--quality",
        action="store_true",
        help="Run pandas data-quality checks on listing sheets",
    )
    parser.add_argument(
        "--fail-on-error",
        action="store_true",
        help="Exit 1 if any check failed",
    )
    parser.add_argument(
        "--config",
        default=None,
        help="Local websites-config.yml path (default: load from R2 sheeel_data/monitor/)",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Log each R2 prefix scanned and every file found",
    )
    parser.add_argument(
        "--require-all-scrapers",
        action="store_true",
        help="Fail when any configured scraper has no file, even if category is off-site",
    )
    parser.add_argument(
        "--categories-registry",
        default=None,
        help="Local categories.json path (default: HyperStack/category_monitor/categories.json)",
    )
    parser.add_argument(
        "--row-count-tolerance",
        type=float,
        default=None,
        help=f"Fraction below/above stats min/max allowed (default: {ROW_COUNT_TOLERANCE})",
    )
    return parser.parse_args()


def load_config_local(path: str) -> dict[str, Any]:
    with open(path, encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def load_config_from_r2(client: Any, bucket: str, key: str) -> dict[str, Any]:
    try:
        raw = download_bytes(client, bucket, key)
        return yaml.safe_load(raw) or {}
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code in ("NoSuchKey", "404"):
            raise FileNotFoundError(f"Config not found at r2://{bucket}/{key}") from exc
        raise


def resolve_config(
    client: Any, bucket: str, args: argparse.Namespace
) -> tuple[dict[str, Any], str]:
    local_path = os.environ.get("WEBSITES_CONFIG") or args.config
    if local_path:
        return load_config_local(local_path), f"local:{local_path}"

    source = f"r2://{bucket}/{CONFIG_R2_KEY}"
    print(f"Loading config from {source}")
    return load_config_from_r2(client, bucket, CONFIG_R2_KEY), source


def flatten_scrapers(config: dict[str, Any]) -> list[dict[str, Any]]:
    scrapers: list[dict[str, Any]] = []
    for project in config.get("projects", []):
        repo = project.get("repo", "")
        for scraper in project.get("scrapers", []):
            scrapers.append({**scraper, "repo": repo})
    return scrapers


def schema_by_scraper(config: dict[str, Any]) -> dict[str, dict[str, Any]]:
    mapping: dict[str, dict[str, Any]] = {}
    for entry in config.get("excel_schema", []):
        mapping[entry["scraper"]] = entry
    return mapping


def normalize_slug(value: str) -> str:
    return value.lower().replace("_", "-").strip("/")


def url_slug_from_scraper(scraper: dict[str, Any]) -> str:
    url = scraper.get("url", "")
    return normalize_slug(url.split("/ar/")[-1].replace(".html", ""))


def load_category_registry_local(path: str) -> dict[str, Any]:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def load_category_registry_from_r2(client: Any, bucket: str, key: str) -> dict[str, Any]:
    try:
        raw = download_bytes(client, bucket, key)
        return json.loads(raw)
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code in ("NoSuchKey", "404"):
            return {}
        raise


def resolve_category_registry(
    client: Any, bucket: str, args: argparse.Namespace
) -> tuple[dict[str, Any], str | None]:
    local_path = os.environ.get("CATEGORIES_REGISTRY") or args.categories_registry
    if local_path:
        if not os.path.isfile(local_path):
            print(f"WARN: categories registry not found at {local_path}", file=sys.stderr)
            return {}, None
        return load_category_registry_local(local_path), f"local:{local_path}"

    if os.path.isfile(DEFAULT_CATEGORIES_REGISTRY):
        return (
            load_category_registry_local(DEFAULT_CATEGORIES_REGISTRY),
            f"local:{DEFAULT_CATEGORIES_REGISTRY}",
        )

    registry = load_category_registry_from_r2(client, bucket, CATEGORIES_REGISTRY_R2_KEY)
    if registry:
        return registry, f"r2://{bucket}/{CATEGORIES_REGISTRY_R2_KEY}"

    print("WARN: no categories registry found; all scrapers treated as on-site", file=sys.stderr)
    return {}, None


def site_status_for_scraper(scraper: dict[str, Any], registry: dict[str, Any]) -> str:
    """Return on_site, off_site, or unknown."""
    categories = registry.get("categories", registry)
    if not categories:
        return "unknown"

    scraper_url = scraper.get("url", "")
    url_slug = url_slug_from_scraper(scraper)
    scraper_slug = normalize_slug(scraper.get("slug", ""))

    for cat in categories.values():
        if cat.get("url") == scraper_url:
            return "on_site"

    registry_slugs = {normalize_slug(slug) for slug in categories}
    if url_slug in registry_slugs or scraper_slug in registry_slugs:
        return "on_site"

    return "off_site"


def build_r2_client() -> Any:
    return boto3.client(
        "s3",
        endpoint_url=os.environ["CF_R2_ENDPOINT_URL"],
        aws_access_key_id=os.environ["CF_R2_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["CF_R2_SECRET_ACCESS_KEY"],
        region_name="us-east-1",
        config=Config(signature_version="s3v4", s3={"addressing_style": "path"}),
    )


def strip_bucket_placeholder(r2_path: str) -> str:
    path = r2_path.strip("/")
    path = re.sub(r"^\{r2_bucket\}/", "", path)
    path = re.sub(r"^\{bucket\}/", "", path)
    return path


def is_date_first_partition(config: dict[str, Any]) -> bool:
    """True when R2 keys use year=/month=/day= before {category} (e.g. sheeel_data)."""
    pattern = config.get("meta", {}).get("r2_partition_pattern", "")
    if not pattern or "{category}" not in pattern:
        return False
    year_pos = pattern.find("year=")
    cat_pos = pattern.find("{category}")
    return year_pos >= 0 and year_pos < cat_pos


def partition_prefix(r2_prefix: str, category: str, day: datetime) -> str:
    """Match scraper layout: {prefix}/year=…/month=…/day=…/{category}/excel-files/"""
    return (
        f"{r2_prefix.strip('/')}/year={day.strftime('%Y')}/month={day.strftime('%m')}"
        f"/day={day.strftime('%d')}/{category.strip('/')}/excel-files/"
    )


def iter_dates(end_date: datetime, days: int) -> list[datetime]:
    return [end_date - timedelta(days=offset) for offset in range(days - 1, -1, -1)]


R2_TYPE_BYTE_FIELDS: tuple[tuple[str, str, str], ...] = (
    ("images", "r2_images_bytes", "r2_daily_images_bytes"),
    ("json", "r2_json_bytes", "r2_daily_json_bytes"),
    ("excel", "r2_excel_bytes", "r2_daily_excel_bytes"),
    ("csv", "r2_csv_bytes", "r2_daily_csv_bytes"),
    ("parquet", "r2_parquet_bytes", "r2_daily_parquet_bytes"),
    ("other", "r2_other_bytes", "r2_daily_other_bytes"),
)

SITE_R2_TYPE_BYTE_FIELDS: tuple[tuple[str, str, str], ...] = (
    ("images", "total_r2_images_bytes", "total_r2_daily_images_bytes"),
    ("json", "total_r2_json_bytes", "total_r2_daily_json_bytes"),
    ("excel", "total_r2_excel_bytes", "total_r2_daily_excel_bytes"),
    ("csv", "total_r2_csv_bytes", "total_r2_daily_csv_bytes"),
    ("parquet", "total_r2_parquet_bytes", "total_r2_daily_parquet_bytes"),
    ("other", "total_r2_other_bytes", "total_r2_daily_other_bytes"),
)


def apply_scraper_r2_type_fields(
    scraper_result: dict[str, Any],
    total_inventory: dict[str, Any],
    daily_inventory: dict[str, Any],
) -> None:
    scraper_result["r2_file_count"] = total_inventory["objects"]
    scraper_result["r2_size_bytes"] = total_inventory["size_bytes"]
    scraper_result["r2_daily_size"] = daily_inventory["size_bytes"]
    for category, total_field, daily_field in R2_TYPE_BYTE_FIELDS:
        scraper_result[total_field] = total_inventory["by_type_bytes"].get(category, 0)
        scraper_result[daily_field] = daily_inventory["by_type_bytes"].get(category, 0)


def list_xlsx_keys(client: Any, bucket: str, prefix: str) -> list[str]:
    keys: list[str] = []
    paginator = client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            if key.lower().endswith(".xlsx"):
                keys.append(key)
    return keys


def download_bytes(client: Any, bucket: str, key: str) -> bytes:
    buf = io.BytesIO()
    client.download_fileobj(bucket, key, buf)
    return buf.getvalue()


def inspect_workbook(data: bytes) -> dict[str, dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets: dict[str, dict[str, Any]] = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(rows, None)
            headers = [str(c).strip() if c is not None else "" for c in (header_row or [])]
            headers = [h for h in headers if h]
            row_count = max(ws.max_row - 1, 0) if ws.max_row else 0
            sheets[name] = {"headers": headers, "row_count": row_count}
    finally:
        wb.close()
    return sheets


def is_pattern_name(name: str) -> bool:
    return name.startswith("{") and name.endswith("}")


def match_sheet_rules(
    sheet_rules: list[dict[str, Any]], actual_sheets: set[str]
) -> list[tuple[dict[str, Any], str | None]]:
    """Return (rule, matched_sheet_name) pairs to validate."""
    matched: list[tuple[dict[str, Any], str | None]] = []
    used_sheets: set[str] = set()

    for rule in sheet_rules:
        expected = rule["name"]
        if is_pattern_name(expected):
            pattern_label = expected.strip("{}")
            candidates = [
                s
                for s in actual_sheets
                if s not in used_sheets and s != "ALL_PRODUCTS"
            ]
            if pattern_label == "subcategory":
                for sheet in sorted(candidates):
                    matched.append((rule, sheet))
                    used_sheets.add(sheet)
            continue
        if expected in actual_sheets:
            matched.append((rule, expected))
            used_sheets.add(expected)
        else:
            matched.append((rule, None))

    return matched


def check_result(name: str, passed: bool, detail: str, severity: str = "critical") -> dict[str, Any]:
    return {
        "check": name,
        "passed": passed,
        "detail": detail,
        "severity": severity,
    }


def normalize_product_id(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def product_ids_from_sheet(data: bytes, sheet_name: str) -> set[str]:
    return set(product_id_counts_from_sheet(data, sheet_name).keys())


def product_id_counts_from_sheet(data: bytes, sheet_name: str) -> dict[str, int]:
    df = pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine="openpyxl")
    if "product_id" not in df.columns:
        return {}
    counts: dict[str, int] = {}
    for value in df["product_id"]:
        normalized = normalize_product_id(value)
        if normalized is not None:
            counts[normalized] = counts.get(normalized, 0) + 1
    return counts


def _sample_ids(ids: set[str] | list[str], limit: int = 5) -> str:
    if not ids:
        return ""
    ordered = sorted(ids)[:limit]
    return f" (sample: {', '.join(ordered)})"


def run_cross_sheet_product_id_checks(data: bytes, sheet_names: list[str]) -> list[dict[str, Any]]:
    """For multi-sheet workbooks: verify ALL_PRODUCTS ids match subcategory sheets."""
    if "ALL_PRODUCTS" not in sheet_names:
        return []

    sub_sheets = sorted(name for name in sheet_names if name != "ALL_PRODUCTS")
    if not sub_sheets:
        return []

    try:
        all_products_ids = product_ids_from_sheet(data, "ALL_PRODUCTS")
    except Exception as exc:
        return [
            check_result(
                "all_products_ids_in_subcategory_sheets",
                False,
                f"failed to read ALL_PRODUCTS: {exc}",
                "critical",
            )
        ]

    subcategory_union: set[str] = set()
    read_errors: list[str] = []
    for sheet in sub_sheets:
        try:
            subcategory_union |= product_ids_from_sheet(data, sheet)
        except Exception as exc:
            read_errors.append(f"{sheet}: {exc}")

    if read_errors:
        return [
            check_result(
                "all_products_ids_in_subcategory_sheets",
                False,
                f"failed to read subcategory sheet(s): {'; '.join(read_errors)}",
                "critical",
            )
        ]

    missing_in_subs = all_products_ids - subcategory_union
    missing_in_all = subcategory_union - all_products_ids

    checks = [
        check_result(
            "all_products_ids_in_subcategory_sheets",
            not missing_in_subs,
            (
                f"{len(missing_in_subs)} id(s) in ALL_PRODUCTS not found in any subcategory sheet"
                f"{_sample_ids(missing_in_subs)}"
            ),
            "critical",
        ),
        check_result(
            "subcategory_ids_in_all_products",
            not missing_in_all,
            (
                f"{len(missing_in_all)} id(s) in subcategory sheets not found in ALL_PRODUCTS"
                f"{_sample_ids(missing_in_all)}"
            ),
            "high",
        ),
        check_result(
            "unique_product_id_set_match",
            all_products_ids == subcategory_union,
            (
                f"unique ALL_PRODUCTS={len(all_products_ids)}, "
                f"union of {len(sub_sheets)} subcategory sheet(s)={len(subcategory_union)}"
            ),
            "high",
        ),
    ]
    return checks


def run_all_products_duplicate_checks(data: bytes, sheet_names: list[str]) -> list[dict[str, Any]]:
    """
    ALL_PRODUCTS row duplicates are OK when the same id appears in 2+ subcategory sheets.
    Duplicates not explained that way are failures.
    """
    if "ALL_PRODUCTS" not in sheet_names:
        return []

    sub_sheets = sorted(name for name in sheet_names if name != "ALL_PRODUCTS")
    if not sub_sheets:
        return []

    try:
        all_counts = product_id_counts_from_sheet(data, "ALL_PRODUCTS")
    except Exception as exc:
        return [
            check_result(
                "duplicate_product_id_unexplained",
                False,
                f"failed to read ALL_PRODUCTS: {exc}",
                "high",
            )
        ]

    duplicate_ids = {pid for pid, count in all_counts.items() if count > 1}
    if not duplicate_ids:
        return []

    id_to_sub_sheets: dict[str, set[str]] = {pid: set() for pid in duplicate_ids}
    read_errors: list[str] = []
    for sheet in sub_sheets:
        try:
            sheet_counts = product_id_counts_from_sheet(data, sheet)
        except Exception as exc:
            read_errors.append(f"{sheet}: {exc}")
            continue
        for pid in duplicate_ids:
            if pid in sheet_counts:
                id_to_sub_sheets[pid].add(sheet)

    if read_errors:
        return [
            check_result(
                "duplicate_product_id_unexplained",
                False,
                f"failed to read subcategory sheet(s): {'; '.join(read_errors)}",
                "high",
            )
        ]

    explained: list[str] = []
    unexplained: list[str] = []
    for pid in duplicate_ids:
        if len(id_to_sub_sheets[pid]) >= 2:
            explained.append(pid)
        else:
            unexplained.append(pid)

    extra_rows = sum(all_counts[pid] - 1 for pid in explained)
    checks: list[dict[str, Any]] = []

    if explained:
        checks.append(
            check_result(
                "duplicate_product_id_multisheet",
                True,
                (
                    f"{len(explained)} id(s), {extra_rows} extra row(s) in ALL_PRODUCTS — "
                    f"product listed in multiple subcategory sheets (expected)"
                    f"{_sample_ids(explained)}"
                ),
                "medium",
            )
        )

    checks.append(
        check_result(
            "duplicate_product_id_unexplained",
            not unexplained,
            (
                f"{len(unexplained)} id(s) duplicated in ALL_PRODUCTS but not in 2+ subcategory sheets"
                f"{_sample_ids(unexplained)}"
            ),
            "high",
        )
    )
    return checks


def resolve_row_count_bounds(
    *,
    scraper_name: str,
    sheet_name: str,
    row_count: int,
    rule: dict[str, Any],
    stats: dict[str, Any],
    tolerance: float,
) -> tuple[int, int, str]:
    """
    Derive row-count bounds from monitor_stats when available.
    Without stats, only enforce schema minimum (no upper cap).
    """
    schema_lo, _schema_hi = rule.get("row_count_range", [1, ROW_COUNT_NO_MAX])
    sheet_stats = (
        stats.get("scrapers", {}).get(scraper_name, {}).get("sheets", {}).get(sheet_name, {})
    )
    obs_min = sheet_stats.get("row_count_min")
    obs_max = sheet_stats.get("row_count_max")

    if obs_min is not None and obs_max is not None:
        lo = max(1, int(obs_min * (1 - tolerance)))
        hi = int(max(obs_max, row_count) * (1 + tolerance))
        return lo, hi, f"stats [{obs_min},{obs_max}] ±{int(tolerance * 100)}%"

    lo = max(1, int(schema_lo)) if schema_lo else 1
    return lo, ROW_COUNT_NO_MAX, "min-only (no stats yet)"


def validate_file(
    *,
    key: str,
    size_kb: float,
    schema: dict[str, Any],
    sheet_data: dict[str, dict[str, Any]],
    inspect_date: datetime,
    scraper_name: str,
    stats: dict[str, Any],
    row_count_tolerance: float,
) -> dict[str, Any]:
    filename = os.path.basename(key)
    checks: list[dict[str, Any]] = []
    actual_sheets = set(sheet_data.keys())

    pattern = schema.get("excel_file_pattern", "*.xlsx")
    checks.append(
        check_result(
            "filename_pattern",
            fnmatch.fnmatch(filename, pattern),
            f"file={filename}, pattern={pattern}",
            "high",
        )
    )

    min_kb = schema.get("min_file_size_kb", 0)
    checks.append(
        check_result(
            "min_file_size_kb",
            size_kb >= min_kb,
            f"size_kb={size_kb:.1f}, min={min_kb}",
            "medium",
        )
    )

    checks.append(
        check_result(
            "file_readable",
            bool(sheet_data),
            "openpyxl read succeeded" if sheet_data else "no sheets found",
            "critical",
        )
    )

    for rule, matched_sheet in match_sheet_rules(schema.get("sheets", []), actual_sheets):
        expected_name = rule["name"]
        if matched_sheet is None:
            if is_pattern_name(expected_name):
                continue
            checks.append(
                check_result(
                    "sheet_exists",
                    False,
                    f"missing sheet '{expected_name}'",
                    "critical",
                )
            )
            continue

        info = sheet_data[matched_sheet]
        headers = set(info["headers"])
        required = rule.get("required_columns", [])
        missing = [col for col in required if col not in headers]
        checks.append(
            check_result(
                "required_columns",
                not missing,
                f"sheet={matched_sheet}, missing={missing or 'none'}",
                "critical",
            )
        )

        row_count = info["row_count"]
        lo, hi, source = resolve_row_count_bounds(
            scraper_name=scraper_name,
            sheet_name=matched_sheet,
            row_count=row_count,
            rule=rule,
            stats=stats,
            tolerance=row_count_tolerance,
        )
        checks.append(
            check_result(
                "row_count_range",
                lo <= row_count <= hi,
                f"sheet={matched_sheet}, rows={row_count}, range=[{lo},{hi}] ({source})",
                "high",
            )
        )

    quality_results: dict[str, Any] = {}

    all_passed = all(c["passed"] for c in checks)

    return {
        "key": key,
        "filename": filename,
        "size_kb": round(size_kb, 2),
        "sheets": {
            name: {"headers": data["headers"], "row_count": data["row_count"]}
            for name, data in sheet_data.items()
        },
        "checks": checks,
        "quality": quality_results,
        "all_passed": all_passed,
    }


def run_quality_on_bytes(
    data: bytes, sheet_name: str, inspect_date: datetime
) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    try:
        df = pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine="openpyxl")
    except Exception as exc:
        return {
            "read_dataframe": {
                "passed": False,
                "detail": str(exc),
                "severity": "high",
            }
        }

    if df.empty:
        return {
            "non_empty": {
                "passed": False,
                "detail": "sheet has zero data rows",
                "severity": "high",
            }
        }

    def null_pct(col: str) -> float | None:
        if col not in df.columns:
            return None
        return float(df[col].isna().mean() * 100)

    if "product_id" in df.columns:
        null_id = null_pct("product_id") or 0
        dupes = int(df["product_id"].duplicated().sum())
        results["null_product_id_pct"] = {
            "passed": null_id <= 5.0,
            "detail": f"{null_id:.1f}% null product_id",
            "severity": "critical",
        }
        # ALL_PRODUCTS is a concat across subcategory sheets; duplicates are expected there.
        if sheet_name != "ALL_PRODUCTS":
            results["duplicate_product_id"] = {
                "passed": dupes == 0,
                "detail": f"{dupes} duplicate product_id values",
                "severity": "high",
            }

    if "name" in df.columns:
        null_name = null_pct("name") or 0
        results["null_name_pct"] = {
            "passed": null_name <= 2.0,
            "detail": f"{null_name:.1f}% null name",
            "severity": "high",
        }

    if "special_price" in df.columns:
        null_price = null_pct("special_price") or 0
        results["null_price_pct"] = {
            "passed": null_price <= 10.0,
            "detail": f"{null_price:.1f}% null special_price",
            "severity": "medium",
        }

    if "scraped_at" in df.columns:
        target = inspect_date.strftime("%Y-%m-%d")
        try:
            dates = pd.to_datetime(df["scraped_at"], errors="coerce").dt.strftime("%Y-%m-%d")
            stale_pct = float((dates != target).mean() * 100)
        except Exception:
            stale_pct = 100.0
        results["stale_scraped_at_pct"] = {
            "passed": stale_pct <= 5.0,
            "detail": f"{stale_pct:.1f}% scraped_at not on {target}",
            "severity": "medium",
        }

    return results


def load_existing_stats(client: Any, bucket: str, key: str) -> dict[str, Any]:
    try:
        obj = client.get_object(Bucket=bucket, Key=key)
        return yaml.safe_load(obj["Body"].read()) or {}
    except ClientError as exc:
        if exc.response["Error"]["Code"] in ("NoSuchKey", "404"):
            return {}
        raise


def merge_stats(
    existing: dict[str, Any],
    scraper_name: str,
    file_results: list[dict[str, Any]],
) -> dict[str, Any]:
    stats = existing.setdefault("scrapers", {})
    entry = stats.setdefault(scraper_name, {"files": 0, "sheets": {}})

    for fr in file_results:
        entry["files"] = entry.get("files", 0) + 1
        size_kb = fr.get("size_kb", 0)
        entry["min_file_size_kb"] = min(entry.get("min_file_size_kb", size_kb), size_kb)
        entry["max_file_size_kb"] = max(entry.get("max_file_size_kb", size_kb), size_kb)

        for sheet_name, sheet_info in fr.get("sheets", {}).items():
            sh = entry["sheets"].setdefault(sheet_name, {"columns": [], "row_count_min": None, "row_count_max": None})
            rows = sheet_info.get("row_count", 0)
            sh["row_count_min"] = rows if sh["row_count_min"] is None else min(sh["row_count_min"], rows)
            sh["row_count_max"] = rows if sh["row_count_max"] is None else max(sh["row_count_max"], rows)
            cols = set(sh.get("columns", []))
            cols.update(sheet_info.get("headers", []))
            sh["columns"] = sorted(cols)

    existing["updated_at"] = datetime.now(timezone.utc).isoformat()
    return existing


def upload_text(client: Any, bucket: str, key: str, body: str, content_type: str) -> None:
    client.put_object(
        Bucket=bucket,
        Key=key,
        Body=body.encode("utf-8"),
        ContentType=content_type,
    )


def write_step_summary(lines: list[str]) -> None:
    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if not summary_path:
        return
    with open(summary_path, "a", encoding="utf-8") as fh:
        fh.write("\n".join(lines))
        fh.write("\n")


def print_summary_table(results: list[dict[str, Any]]) -> None:
    print(f"\n{'Scraper':<35} {'Site':>4} {'Files':>5} {'Pass':>5} {'Total':>5} {'OK':>4}")
    print("-" * 63)
    for row in results:
        site = {"on_site": "ON", "off_site": "OFF", "unknown": "?"}.get(
            row.get("site_status", "unknown"), "?"
        )
        if row.get("monitor_status") == "skipped_off_site":
            ok = "SKIP"
        else:
            ok = "YES" if row["all_passed"] else "NO"
        print(
            f"{row['scraper']:<35} {site:>4} {row['files_found']:>5} "
            f"{row['checks_passed']:>5} {row['checks_total']:>5} {ok:>4}"
        )


def failed_checks_for_scraper(scraper_result: dict[str, Any]) -> list[dict[str, Any]]:
    failures: list[dict[str, Any]] = []
    for file_result in scraper_result.get("files", []):
        source = file_result.get("filename") or file_result.get("key") or "(no file)"
        for chk in file_result.get("checks", []):
            if not chk.get("passed"):
                failures.append({**chk, "source": source})
        for check_name, quality in file_result.get("quality", {}).items():
            if not quality.get("passed", True):
                failures.append(
                    {
                        "check": check_name,
                        "passed": False,
                        "detail": quality.get("detail", ""),
                        "severity": quality.get("severity", "medium"),
                        "source": source,
                    }
                )
    return failures


def alarms_for_scraper(scraper_result: dict[str, Any]) -> list[dict[str, Any]]:
    alarms: list[dict[str, Any]] = []
    for file_result in scraper_result.get("files", []):
        source = file_result.get("filename") or file_result.get("key") or "(no file)"
        for chk in file_result.get("checks", []):
            if chk.get("passed") and chk.get("severity") == "medium" and chk.get("check", "").startswith(
                "duplicate_product_id"
            ):
                alarms.append({**chk, "source": source})
    return alarms


def print_alarm_details(results: list[dict[str, Any]]) -> None:
    alarm_rows: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for row in results:
        for alarm in alarms_for_scraper(row):
            alarm_rows.append((row, alarm))

    if not alarm_rows:
        return

    print(f"\n{'=' * 63}")
    print(f"Warnings ({len(alarm_rows)} — passed, informational)")
    print(f"{'=' * 63}")
    for row, alarm in alarm_rows:
        source = alarm.get("source", "")
        source_suffix = f"  ({source})" if source and source != "(no file)" else ""
        print(
            f"  WARN [{alarm.get('severity', 'medium')}] {row['scraper']}: "
            f"{alarm.get('check')} — {alarm.get('detail')}{source_suffix}"
        )


def print_run_header(
    *,
    bucket: str,
    r2_prefix: str,
    dates: list[datetime],
    args: argparse.Namespace,
    registry_source: str | None,
    on_site_count: int,
    off_site_count: int,
) -> None:
    date_labels = ", ".join(d.strftime("%Y-%m-%d") for d in dates)
    print(f"Bucket: r2://{bucket}/")
    print(f"Inspect dates ({args.days_lookback} day(s)): {date_labels}")
    print(f"R2 prefix: {r2_prefix}/year=…/month=…/day=…/{{slug}}/excel-files/")
    if registry_source:
        print(
            f"Category registry: {registry_source} "
            f"({on_site_count} on-site, {off_site_count} off-site scrapers)"
        )
    else:
        print("Category registry: not loaded (all scrapers treated as on-site)")
    flags = []
    if args.quality:
        flags.append("quality")
    if args.update_stats:
        flags.append("update-stats")
    if args.verbose:
        flags.append("verbose")
    if args.require_all_scrapers:
        flags.append("require-all-scrapers")
    if flags:
        print(f"Options: {', '.join(flags)}")


def print_failure_details(
    results: list[dict[str, Any]],
    *,
    bucket: str,
) -> None:
    skipped_rows = [row for row in results if row.get("monitor_status") == "skipped_off_site"]
    failed_rows = [
        row
        for row in results
        if not row["all_passed"] and row.get("monitor_status") != "skipped_off_site"
    ]

    if skipped_rows:
        print(f"\n{'=' * 63}")
        print(f"Off-site categories skipped ({len(skipped_rows)} scraper(s), no file expected)")
        print(f"{'=' * 63}")
        for row in skipped_rows:
            url_slug = url_slug_from_scraper({"url": row.get("url", "")})
            print(
                f"  {row['scraper']} [{row.get('slug')}] — "
                f"not on sheeel.com menu ({url_slug or 'unknown url'})"
            )

    if not failed_rows:
        if skipped_rows:
            print("\nAll on-site scrapers passed.")
        else:
            print("\nAll scrapers passed.")
        return

    print(f"\n{'=' * 63}")
    print(f"Failure details ({len(failed_rows)} on-site scraper(s))")
    print(f"{'=' * 63}")

    for row in failed_rows:
        name = row["scraper"]
        slug = row.get("slug") or "?"
        site_note = row.get("site_status", "unknown")
        print(
            f"\n{name} [{slug}] ({site_note}) — "
            f"{row['files_found']} file(s), {row['checks_passed']}/{row['checks_total']} checks passed"
        )

        if row["files_found"] == 0:
            for prefix in row.get("searched_prefixes", []):
                print(f"  scanned: r2://{bucket}/{prefix}  →  0 .xlsx")
            print("  → category is on-site but scraper did not upload a file")

        for failure in failed_checks_for_scraper(row):
            severity = failure.get("severity", "?")
            check = failure.get("check", "?")
            detail = failure.get("detail", "")
            source = failure.get("source", "")
            source_suffix = f"  ({source})" if source and source != "(no file)" else ""
            print(f"  FAIL [{severity}] {check}: {detail}{source_suffix}")


def main() -> int:
    args = parse_args()
    run_started_at = datetime.now(timezone.utc)

    bucket = os.environ.get("CF_R2_BUCKET_NAME")
    if not bucket:
        print("ERROR: CF_R2_BUCKET_NAME not set", file=sys.stderr)
        return 1

    client = build_r2_client()

    try:
        config, config_source = resolve_config(client, bucket, args)
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    scrapers = flatten_scrapers(config)
    schemas = schema_by_scraper(config)
    category_registry, registry_source = resolve_category_registry(client, bucket, args)

    bucket = bucket or config.get("meta", {}).get("r2_bucket")
    r2_prefix = config.get("meta", {}).get("r2_prefix", DEFAULT_R2_PREFIX).strip("/")

    on_site_count = sum(
        1 for s in scrapers if site_status_for_scraper(s, category_registry) == "on_site"
    )
    off_site_count = sum(
        1 for s in scrapers if site_status_for_scraper(s, category_registry) == "off_site"
    )

    if args.date:
        end_date = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    else:
        end_date = datetime.now(timezone.utc) - timedelta(days=1)

    dates = iter_dates(end_date, args.days_lookback)
    report: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "config_source": config_source,
        "category_registry_source": registry_source,
        "inspect_dates": [d.strftime("%Y-%m-%d") for d in dates],
        "scrapers": [],
    }

    summary_rows: list[dict[str, Any]] = []
    any_failure = False
    stats_key = f"{r2_prefix}/{MONITOR_SUBPATH}/monitor_stats.yml"
    stats_blob: dict[str, Any] = load_existing_stats(client, bucket, stats_key)
    row_count_tolerance = (
        args.row_count_tolerance if args.row_count_tolerance is not None else ROW_COUNT_TOLERANCE
    )

    print_run_header(
        bucket=bucket,
        r2_prefix=r2_prefix,
        dates=dates,
        args=args,
        registry_source=registry_source,
        on_site_count=on_site_count,
        off_site_count=off_site_count,
    )

    date_first_layout = is_date_first_partition(config)

    for scraper in scrapers:
        name = scraper["name"]
        schema = schemas.get(name)
        if not schema:
            print(f"WARN: no excel_schema for scraper '{name}'", file=sys.stderr)
            continue

        category = scraper.get("slug") or strip_bucket_placeholder(scraper["r2_path"]).rsplit("/", 1)[-1]
        site_status = site_status_for_scraper(scraper, category_registry)
        scraper_result: dict[str, Any] = {
            "scraper": name,
            "slug": scraper.get("slug"),
            "url": scraper.get("url"),
            "repo": scraper.get("repo"),
            "site_status": site_status,
            "monitor_status": "pending",
            "searched_prefixes": [],
            "files": [],
            "files_found": 0,
            "checks_passed": 0,
            "checks_total": 0,
            "all_passed": True,
        }

        file_results_for_stats: list[dict[str, Any]] = []
        excel_downloads: list[tuple[str, bytes]] = []

        for day in dates:
            prefix = partition_prefix(r2_prefix, category, day)
            scraper_result["searched_prefixes"].append(prefix)
            keys = list_xlsx_keys(client, bucket, prefix)

            if args.verbose or (not keys and site_status != "off_site"):
                day_label = day.strftime("%Y-%m-%d")
                skip_note = " (off-site, skipped)" if not keys and site_status == "off_site" else ""
                print(
                    f"  {name}: r2://{bucket}/{prefix}  "
                    f"({day_label})  →  {len(keys)} .xlsx{skip_note}"
                )
                if args.verbose and keys:
                    for key in keys:
                        print(f"    found: {os.path.basename(key)}")

            for key in keys:
                try:
                    head = client.head_object(Bucket=bucket, Key=key)
                    size_kb = head["ContentLength"] / 1024
                    raw = download_bytes(client, bucket, key)
                    excel_downloads.append((key, raw))
                    sheet_data = inspect_workbook(raw)

                    file_result = validate_file(
                        key=key,
                        size_kb=size_kb,
                        schema=schema,
                        sheet_data=sheet_data,
                        inspect_date=day,
                        scraper_name=name,
                        stats=stats_blob,
                        row_count_tolerance=row_count_tolerance,
                    )

                    if "ALL_PRODUCTS" in sheet_data and len(sheet_data) > 1:
                        sheet_names = list(sheet_data.keys())
                        for cross_chk in run_cross_sheet_product_id_checks(raw, sheet_names):
                            file_result["checks"].append(cross_chk)
                            if not cross_chk["passed"]:
                                file_result["all_passed"] = False
                        for dup_chk in run_all_products_duplicate_checks(raw, sheet_names):
                            file_result["checks"].append(dup_chk)
                            if not dup_chk["passed"]:
                                file_result["all_passed"] = False

                    if args.quality:
                        target = (
                            "ALL_PRODUCTS"
                            if "ALL_PRODUCTS" in sheet_data
                            else next(iter(sheet_data), "Sheet1")
                        )
                        file_result["quality"] = run_quality_on_bytes(raw, target, day)
                        q_ok = all(q.get("passed", True) for q in file_result["quality"].values())
                        file_result["all_passed"] = file_result["all_passed"] and q_ok

                    scraper_result["files"].append(file_result)
                    file_results_for_stats.append(file_result)
                    scraper_result["files_found"] += 1

                    for chk in file_result["checks"]:
                        scraper_result["checks_total"] += 1
                        if chk["passed"]:
                            scraper_result["checks_passed"] += 1
                        else:
                            scraper_result["all_passed"] = False
                            any_failure = True

                    for q in file_result.get("quality", {}).values():
                        scraper_result["checks_total"] += 1
                        if q.get("passed", True):
                            scraper_result["checks_passed"] += 1
                        else:
                            scraper_result["all_passed"] = False
                            any_failure = True

                except Exception as exc:
                    any_failure = True
                    scraper_result["all_passed"] = False
                    scraper_result["files_found"] += 1
                    scraper_result["checks_total"] += 1
                    scraper_result["files"].append(
                        {
                            "key": key,
                            "error": str(exc),
                            "all_passed": False,
                            "checks": [
                                check_result("file_processing", False, str(exc), "critical")
                            ],
                        }
                    )

        if scraper_result["files_found"] == 0:
            if site_status == "off_site" and not args.require_all_scrapers:
                scraper_result["monitor_status"] = "skipped_off_site"
                scraper_result["all_passed"] = True
            else:
                scraper_result["monitor_status"] = "failed"
                scraper_result["all_passed"] = False
                any_failure = True
                scraper_result["checks_total"] += 1
                scraper_result["files"].append(
                    {
                        "checks": [
                            check_result(
                                "files_found",
                                False,
                                (
                                    f"no .xlsx for {dates[-1].strftime('%Y-%m-%d')}; "
                                    f"searched r2://{bucket}/{scraper_result['searched_prefixes'][-1]}"
                                ),
                                "critical",
                            )
                        ],
                        "all_passed": False,
                    }
                )
        elif scraper_result["all_passed"]:
            scraper_result["monitor_status"] = "passed"
        else:
            scraper_result["monitor_status"] = "failed"

        if args.update_stats and file_results_for_stats:
            stats_blob = merge_stats(stats_blob, name, file_results_for_stats)

        r2_base = partition_prefix(r2_prefix, category, end_date).replace("/excel-files/", "/")
        try:
            ads_stats = count_scraper_ads(client, bucket, r2_base, end_date, excel_downloads)
        except Exception as exc:
            print(f"WARN: ads count failed for '{name}': {exc}", file=sys.stderr)
            ads_stats = {"unique_ads": 0, "total_rows": 0, "ads_source": "none"}
        scraper_result["unique_ads"] = ads_stats.get("unique_ads") or 0
        scraper_result["total_rows"] = ads_stats.get("total_rows") or 0
        scraper_result["ads_source"] = ads_stats.get("ads_source", "none")

        try:
            req_stats = count_scraper_request_metrics(client, bucket, r2_base, end_date)
        except Exception as exc:
            print(f"WARN: request metrics failed for '{name}': {exc}", file=sys.stderr)
            req_stats = {"metrics_source": "none"}
        scraper_result["requests_total"] = req_stats.get("requests_total")
        scraper_result["requests_failed"] = req_stats.get("requests_failed")
        scraper_result["error_rate_pct"] = req_stats.get("error_rate_pct")
        scraper_result["requests_per_min"] = req_stats.get("requests_per_min")
        scraper_result["duration_sec"] = req_stats.get("duration_sec")
        scraper_result["metrics_source"] = req_stats.get("metrics_source", "none")
        if req_stats.get("failed_items_summary"):
            scraper_result["failed_items_summary"] = req_stats["failed_items_summary"]

        if date_first_layout:
            inventory_base = r2_prefix
            inventory_path_contains = f"/{category}/"
        else:
            inventory_base = strip_bucket_placeholder(scraper.get("r2_path", "")).strip("/")
            inventory_path_contains = None

        print(f"  {name}: counting R2 inventory under {inventory_base}/...")
        try:
            total_inventory = count_scraper_r2_inventory_by_type(
                client,
                bucket,
                inventory_base,
                path_contains=inventory_path_contains,
            )
            daily_inventories = [
                count_daily_r2_inventory_by_type(
                    client,
                    bucket,
                    inventory_base,
                    day,
                    path_contains=inventory_path_contains,
                )
                for day in dates
            ]
            daily_inventory = merge_r2_type_inventories(*daily_inventories)
            apply_scraper_r2_type_fields(scraper_result, total_inventory, daily_inventory)
        except Exception as exc:
            print(f"WARN: R2 inventory failed for '{name}': {exc}", file=sys.stderr)
            (
                scraper_result["r2_file_count"],
                scraper_result["r2_size_bytes"],
            ) = count_scraper_r2_inventory(
                client,
                bucket,
                inventory_base,
                path_contains=inventory_path_contains,
            )
            daily_size = 0
            for day in dates:
                day_inventory = count_daily_r2_inventory_by_type(
                    client,
                    bucket,
                    inventory_base,
                    day,
                    path_contains=inventory_path_contains,
                )
                daily_size += day_inventory["size_bytes"]
            scraper_result["r2_daily_size"] = daily_size
            for _, total_field, daily_field in R2_TYPE_BYTE_FIELDS:
                scraper_result[total_field] = 0
                scraper_result[daily_field] = 0

        report["scrapers"].append(scraper_result)
        summary_rows.append(scraper_result)

    report["total_unique_ads"] = sum(
        r.get("unique_ads") or 0 for r in report["scrapers"]
    )

    site_r2_prefix = config.get("meta", {}).get("r2_prefix", "").strip("/")
    if site_r2_prefix:
        print(f"Counting site R2 inventory under {site_r2_prefix}/...")
        try:
            site_inventory = count_site_r2_inventory_by_type(client, bucket, site_r2_prefix)
            report["total_r2_files"] = site_inventory["objects"]
            report["total_r2_size_bytes"] = site_inventory["size_bytes"]
            for category, total_field, _ in SITE_R2_TYPE_BYTE_FIELDS:
                report[total_field] = site_inventory["by_type_bytes"].get(category, 0)
        except Exception as exc:
            print(f"WARN: site R2 inventory failed: {exc}", file=sys.stderr)
            (
                report["total_r2_files"],
                report["total_r2_size_bytes"],
            ) = count_site_r2_inventory(client, bucket, site_r2_prefix)
            for _, total_field, _ in SITE_R2_TYPE_BYTE_FIELDS:
                report[total_field] = 0
    else:
        report["total_r2_files"] = sum(
            r.get("r2_file_count") or 0 for r in report["scrapers"]
        )
        report["total_r2_size_bytes"] = sum(
            r.get("r2_size_bytes") or 0 for r in report["scrapers"]
        )
        for (_, scraper_total, _), (_, site_total, _) in zip(
            R2_TYPE_BYTE_FIELDS,
            SITE_R2_TYPE_BYTE_FIELDS,
        ):
            report[site_total] = sum(r.get(scraper_total, 0) or 0 for r in report["scrapers"])

    report["total_r2_daily_size"] = sum(
        r.get("r2_daily_size") or 0 for r in report["scrapers"]
    )
    for (_, _, scraper_daily), (_, _, site_daily) in zip(
        R2_TYPE_BYTE_FIELDS,
        SITE_R2_TYPE_BYTE_FIELDS,
    ):
        report[site_daily] = sum(r.get(scraper_daily, 0) or 0 for r in report["scrapers"])

    site_metrics = aggregate_site_request_metrics(report["scrapers"])
    report.update(site_metrics)
    report["error_summary"] = build_run_error_summary(report["scrapers"], [])

    report_date = end_date.strftime("%Y-%m-%d")
    passed_n = sum(1 for r in summary_rows if r.get("monitor_status") == "passed")
    failed_n = sum(1 for r in summary_rows if r.get("monitor_status") == "failed")

    site_meta = load_site_run_meta(client=client, bucket=bucket)
    github_run = build_scraper_run_meta(
        site_meta,
        report_date,
        run_started_at.replace(tzinfo=None),
        failed_n == 0,
        summary_rows,
    )
    github_gmail = (site_meta.get("github_gmail") or site_meta.get("github_email") or "").strip()
    if github_gmail:
        github_run["github_gmail"] = github_gmail
    report["github_run"] = github_run
    if github_gmail:
        report["github_gmail"] = github_gmail
    report["run_place"] = report["github_run"].get("run_place")

    report_key = f"{r2_prefix}/{MONITOR_SUBPATH}/{report_date}/report.json"
    upload_text(client, bucket, report_key, json.dumps(report, indent=2), "application/json")

    if args.update_stats:
        upload_text(client, bucket, stats_key, yaml.dump(stats_blob, sort_keys=False), "text/yaml")

    print_summary_table(summary_rows)
    print_alarm_details(summary_rows)
    print_failure_details(summary_rows, bucket=bucket)

    skipped_n = sum(1 for r in summary_rows if r.get("monitor_status") == "skipped_off_site")
    print(
        f"\nRun summary: {passed_n} passed, {failed_n} failed (on-site), "
        f"{skipped_n} skipped (off-site)"
    )
    print(f"Report uploaded: r2://{bucket}/{report_key}")

    summary_lines = [
        "## R2 Excel Schema Monitor",
        "",
        f"Dates: {', '.join(d.strftime('%Y-%m-%d') for d in dates)}",
        f"On-site: {on_site_count} | Off-site: {off_site_count} | "
        f"Passed: {passed_n} | Failed: {failed_n} | Skipped: {skipped_n}",
        "",
        "| Scraper | Site | Files | Passed | Total | OK |",
        "| --- | --- | ---: | ---: | ---: | --- |",
    ]
    for row in summary_rows:
        site = {"on_site": "ON", "off_site": "OFF", "unknown": "?"}.get(
            row.get("site_status", "unknown"), "?"
        )
        if row.get("monitor_status") == "skipped_off_site":
            ok = "⏭️"
        else:
            ok = "✅" if row["all_passed"] else "❌"
        summary_lines.append(
            f"| {row['scraper']} | {site} | {row['files_found']} | {row['checks_passed']} | "
            f"{row['checks_total']} | {ok} |"
        )

    skipped_rows = [row for row in summary_rows if row.get("monitor_status") == "skipped_off_site"]
    if skipped_rows:
        summary_lines.extend(["", "### Skipped (off-site)", ""])
        for row in skipped_rows:
            summary_lines.append(f"- ⏭️ **{row['scraper']}** — not on sheeel.com menu")

    failed_rows = [
        row
        for row in summary_rows
        if not row["all_passed"] and row.get("monitor_status") != "skipped_off_site"
    ]
    if failed_rows:
        summary_lines.extend(["", "### Failures (on-site)", ""])
        for row in failed_rows:
            summary_lines.append(
                f"**{row['scraper']}** — {row['checks_passed']}/{row['checks_total']} passed, "
                f"{row['files_found']} file(s)"
            )
            if row["files_found"] == 0:
                summary_lines.append("- category on-site but no file uploaded")
                for prefix in row.get("searched_prefixes", []):
                    summary_lines.append(f"- scanned `r2://{bucket}/{prefix}` → 0 .xlsx")
            for failure in failed_checks_for_scraper(row):
                summary_lines.append(
                    f"- ❌ `{failure.get('check')}`: {failure.get('detail')}"
                )
            summary_lines.append("")

    summary_lines.append("")
    summary_lines.append(f"Report: `r2://{bucket}/{report_key}`")
    write_step_summary(summary_lines)

    if args.fail_on_error and any_failure:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
